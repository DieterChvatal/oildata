import pandas as pd
from pprint import pprint as pp
import helpers
import logging
import os
import sys
import time
import config
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from itertools import islice

def summary(filename):
    t0 = time.time()
    fields = ['Unique Code','General Description','Oil/Fluid Long Name','Unit Hours','Oil/Fluid Hours'] # auf die benötigten Felder beschränken

    # # load data into dataframe with openpyxl
    # wb = load_workbook(filename)
    # wbsheets = wb.sheetnames
    # ws0 = wb[wbsheets[0]]
    # data = ws0.values
    # cols = next(data)[1:]
    # data = list(data)
    # idx = [r[0] for r in data]
    # data = (islice(r, 1, None) for r in data)
    # xdf = pd.DataFrame(data, index=idx, columns=cols)[fields]

    xdf = pd.read_excel(filename)[fields]
    logging.info(f"{filename} read into xdf, restricted to [fields] -> next applymap(corr)")
    all_lines = xdf.shape[0]

    corr_xdf = xdf.applymap(helpers.corr) # correct the contents
    logging.info(f"xdf.applymap(helpers.corr) completed -> next start collecting results")
    
    result = {
    'Zeilen': [],
    'Eindeutige Motoren': [],
    'gültige Motoren': [],
    'Kumulierte Öl Stunden': [],
    'Mittlere Öl Stunden pro gültigem Motor': [],
    }
    # alle Öle im file ermitteln
    corr_xdf['Oil_Unique_Name'] = corr_xdf['Oil/Fluid Long Name'].apply(lambda x: str(x).upper().replace(' ',''))
    uniques = list(corr_xdf['Oil_Unique_Name'].unique())
    columns = []
    for oil_name in uniques:
        #new_xdf = corr_xdf[corr_xdf['Oil/Fluid Long Name'] == oil_name] # nur die Zeilen mit 'oil_name' ausfiltern
        new_xdf = corr_xdf[corr_xdf['Oil_Unique_Name'] == oil_name] # nur die Zeilen mit 'oil_name' ausfiltern
        columns.append(new_xdf['Oil/Fluid Long Name'].iloc(0)[0])
        uniquecodes = new_xdf['Unique Code'].unique() # Alle Unique Codes der Motoren auslesen - nur unterschiedliche codes kommen in die Liste
        No_of_Engines = len(uniquecodes)
        result['Zeilen'].append(new_xdf.shape[0])
        result['Eindeutige Motoren'].append(No_of_Engines)
        if new_xdf.shape[0] == 0: #no enties found
            logging.info(f"short before raising Value Error: kein Öl mit Namen '{oil_name}' in '{os.path.basename(filename)}' gefunden.")
            raise ValueError(f"kein Öl mit Namen '{oil_name}' in '{os.path.basename(filename)}' gefunden.")
        oil_sum = 0
        count_sum = 0
        for u in uniquecodes:
            odf = new_xdf[new_xdf['Unique Code'] == u]
            #print(oil_name)
            #print(odf)
            df = odf.sort_values(by = ['Unit Hours'],ascending=[True])
            try: #try to calculate oil running hours
                oil_age = df['Unit Hours'].max() - df['Unit Hours'].min() + df.iloc[0]['Oil/Fluid Hours']
                if oil_age == oil_age: #this is a trick to check for NAN
                    oil_sum += oil_age
                    count_sum += 1
            except Exception:
                logging.info(f"{oil_name}, line with unique code {u}, oil age {oil_age} could not be evaluated, skipping line")
                pass # do nothing on entry rows with missing parameters.
        result['gültige Motoren'].append(count_sum)
        result['Kumulierte Öl Stunden'].append(int(f"{oil_sum:0.0f}"))
        result['Mittlere Öl Stunden pro gültigem Motor'].append(int(f"{oil_sum / (count_sum or 1):0.0f}"))
    result[' '] = ['']*len(columns)
    result['Datum'] = [pd.Timestamp.now()] + ['']* (len(columns)-1)
    result['Datei'] = [os.path.basename(filename)]  + ['']* (len(columns)-1)
    result['Alle Zeilen'] = [all_lines] + ['']* (len(columns)-1)
    rdf = pd.DataFrame.from_dict(result, columns=columns, orient='index')
    logging.info(f"Exporting Zusammenfassung to {config.zoutfile}.")
    #rdf.to_excel(config.zoutfile)
    print(rdf)
    t1 = time.time()
    logging.info(f"Result will now be saved in sheet 'Zusammenfassung' in workbook {filename}")
    # add summary sheets
    wb = load_workbook(filename)
    wb.create_sheet('Zusammenfassung')
    wbsheets = wb.sheetnames
    ws = wb[wbsheets[-1]]
    for r in dataframe_to_rows(rdf, index=True, header=True):
        ws.append(r)
    wb.save(filename)
    wb.close()
    t2 = time.time()

    logging.info(f"Results saved, Dauer Einlesen ...{(t1-t0):0.2f} Dauer Speichern ...{(t2-t1):0.2f} Dauer gesamt ...{(t2-t0):0.2f}")
    #print(f"fertig, Dauer Einlesen ...{(t1-t0):0.2f} Dauer Speichern ...{(t2-t1):0.2f} Dauer gesamt ...{(t2-t0):0.2f}")
    if sys.platform == 'win32':
        logging.info(f"Running on Windows, starting Ecxel {filename}")
        os.startfile(filename)
    else:
        logging.info(f"not Running on Windows.")
        logging.info(rdf)
